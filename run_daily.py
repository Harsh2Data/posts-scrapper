#!/usr/bin/env python3
"""
Daily LinkedIn recruitment-lead pipeline — Steps 4 to 10 of the plan, automated.

POST SOURCE (Step 3) is kept SAFE. This pipeline never logs into or scrapes
LinkedIn. You feed it post text one of two ways:

  • Drop posts you copied while browsing LinkedIn into the ./inbox folder as
    .txt files (one post per file). Works today, no API keys needed.
  • Use linkedin_post_finder.py (Google Custom Search) to auto-discover NEW
    public post URLs, open them, and copy the text into ./inbox.

What this runs automatically every time:
  4  de-duplicate against the master database (by email + post URL)
  5  extract author, company, profile URL, post URL, date
  6  extract all public emails from the post text
  7  validate: valid email + staffing/recruitment relevance + not duplicate
  8  store qualified leads in leads_master.csv
  9  export daily_leads_YYYY_MM_DD.xlsx with the agreed columns
  10 print + save a daily summary report

Run:
  pip install openpyxl --break-system-packages
  python run_daily.py
"""
from __future__ import annotations
import os, re, sys, csv, glob, shutil
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from linkedin_post_extractor import parse_linkedin_post  # reuse the parser

# ---------- config ----------
INBOX      = "inbox"
PROCESSED  = os.path.join(INBOX, "processed")
MASTER     = "leads_master.csv"
MASTER_COLS = ["Contact Person", "Company Name", "Email",
               "LinkedIn Profile", "LinkedIn Post", "Keyword", "Date"]

# Step 2 keywords (used to tag which keyword matched + relevance)
KEYWORDS = ["IT Staffing", "Contract Staffing", "Vendor Collaboration",
            "Recruitment Vendor", "Hiring Partner", "Bench Sales", "C2H",
            "Talent Acquisition", "Immediate Requirement", "Vendor Onboarding"]

# Step 7 relevance gate
RELEVANCE   = re.compile(r"staffing|recruit|vendor|bench|c2c|c2h|corp\s*to\s*corp|"
                         r"contract|talent acquisition|hiring partner|consultant", re.I)
EMAIL_VALID = re.compile(r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$")
POST_URL_RE = re.compile(r"https?://[^\s)]*linkedin\.com/(?:posts|feed/update)/[^\s)]+", re.I)

# If True, store one lead per post (the best/primary email only).
# If False, store every unique email found (honors "extract all emails").
ONE_LEAD_PER_POST = False


def load_master():
    if not os.path.exists(MASTER):
        return []
    with open(MASTER, encoding="utf-8") as f:
        return list(csv.DictReader(f))


LINK_COLS = {"LinkedIn Profile", "LinkedIn Post"}

def write_xlsx(path, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    LINK_FONT  = Font(color="0563C1", underline="single")
    HEADER_FONT = Font(bold=True)
    HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")

    wb = Workbook(); ws = wb.active; ws.title = "Leads"
    ws.append(MASTER_COLS)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([r.get(c, "") for c in MASTER_COLS])
        row_num = ws.max_row
        for col_idx, col_name in enumerate(MASTER_COLS, start=1):
            if col_name in LINK_COLS:
                cell = ws.cell(row=row_num, column=col_idx)
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.font = LINK_FONT

    for i, wd in enumerate([20, 25, 32, 44, 44, 20, 12], start=1):
        ws.column_dimensions[chr(64 + i)].width = wd
    ws.freeze_panes = "A2"
    wb.save(path)


def main(status_cb=None):
    def emit(msg):
        print(msg)
        if status_cb:
            try:
                status_cb(msg)
            except Exception:
                pass

    os.makedirs(PROCESSED, exist_ok=True)
    master = load_master()
    seen_emails = {r["Email"].lower() for r in master if r.get("Email")}
    seen_posts  = {r["LinkedIn Post"].lower() for r in master if r.get("LinkedIn Post")}

    seen_profiles = {r["LinkedIn Profile"].lower() for r in master if r.get("LinkedIn Profile")}
    stats = dict(scanned=0, new_leads=0, duplicates=0, emails=0, no_email=0, irrelevant=0)
    todays = []

    files = sorted(glob.glob(os.path.join(INBOX, "*.txt")))
    emit(f"Extracting leads from {len(files)} scraped post(s)...")

    for path in files:
        text = open(path, encoding="utf-8").read()
        if not text.strip():
            continue
        stats["scanned"] += 1
        parsed = parse_linkedin_post(text)

        emails = [e for e in parsed["all_emails"].split("; ") if e and EMAIL_VALID.match(e)]

        # Step 6 / 7: no email -> skip
        if not emails:
            stats["no_email"] += 1
            shutil.move(path, os.path.join(PROCESSED, os.path.basename(path)))
            continue
        # Step 7: relevance
        if not RELEVANCE.search(text):
            stats["irrelevant"] += 1
            shutil.move(path, os.path.join(PROCESSED, os.path.basename(path)))
            continue

        keyword  = next((k for k in KEYWORDS if k.lower() in text.lower()), "")
        m        = POST_URL_RE.search(text)
        post_url = m.group(0).split("?")[0] if m else ""

        # One row per person — dedup by LinkedIn profile URL first, then primary email
        profile_key = parsed["linkedin_url"].lower() if parsed["linkedin_url"] else None
        primary     = parsed["primary_email"] or emails[0]
        email_key   = primary.lower()

        stats["emails"] += len(emails)
        if (profile_key and profile_key in seen_profiles) or email_key in seen_emails:
            stats["duplicates"] += 1
            emit(f"  - duplicate skipped: {parsed['contact_person']} ({primary})")
            shutil.move(path, os.path.join(PROCESSED, os.path.basename(path)))
            continue

        if profile_key:
            seen_profiles.add(profile_key)
        seen_emails.add(email_key)
        if post_url:
            seen_posts.add(post_url.lower())

        row = {
            "Contact Person":   parsed["contact_person"],
            "Company Name":     parsed["company"],
            "Email":            primary,   # author's primary email only
            "LinkedIn Profile": parsed["linkedin_url"],
            "LinkedIn Post":    post_url,
            "Keyword":          keyword,
            "Date":             datetime.now().strftime("%Y-%m-%d"),
        }
        todays.append(row)
        master.append(row)
        stats["new_leads"] += 1
        emit(f"  + new lead: {parsed['contact_person']} | {parsed['company']} | {primary}")

        shutil.move(path, os.path.join(PROCESSED, os.path.basename(path)))

    # Step 8: persist master
    with open(MASTER, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=MASTER_COLS)
        w.writeheader(); w.writerows(master)

    # Step 9a: master Excel — all-time accumulation, rebuilt every run
    master_xlsx = "leads_master.xlsx"
    write_xlsx(master_xlsx, master)

    # Step 9b: daily Excel — only today's new leads
    daily_xlsx = f"daily_leads_{datetime.now():%Y_%m_%d}.xlsx"
    write_xlsx(daily_xlsx, todays)

    # Step 10: report
    report = (
        f"\n===== LEAD REPORT  {datetime.now():%Y-%m-%d %H:%M} =====\n"
        f"  Posts scanned        : {stats['scanned']}\n"
        f"  New leads added      : {stats['new_leads']}\n"
        f"  Duplicates skipped   : {stats['duplicates']}\n"
        f"  Emails found         : {stats['emails']}\n"
        f"  Posts with no email  : {stats['no_email']}\n"
        f"  Off-topic skipped    : {stats['irrelevant']}\n"
        f"  Master Excel         : {master_xlsx}  (all-time, {len(master)} leads)\n"
        f"  Today's Excel        : {daily_xlsx}  ({len(todays)} new today)\n"
        f"  Master database      : {MASTER}  ({len(master)} total leads)\n"
        f"================================================\n"
    )
    emit(report)
    with open("report_latest.txt", "w", encoding="utf-8") as f:
        f.write(report)
    return {
        "master_xlsx": master_xlsx,
        "daily_xlsx": daily_xlsx,
        "stats": stats,
    }


if __name__ == "__main__":
    main()
