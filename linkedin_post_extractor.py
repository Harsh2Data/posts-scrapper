#!/usr/bin/env python3
"""
LinkedIn post -> lead extractor.

Paste a copied LinkedIn post (the kind where someone advertises that they're
looking for IT contract-staffing / vendor partners and drops their email),
and this turns it into a clean structured lead row:

    contact_person, title, company, primary_email, all_emails,
    phone, service_type, linkedin_url, company_url, post_age, hashtags

It then appends the row to an Excel file (and/or a Google Sheet) with
de-duplication, so the same person never gets added twice.

IMPORTANT — this tool only PARSES text you give it.
You still find the posts yourself by browsing/searching LinkedIn normally in
your own account (which is allowed). Nothing here logs into or scrapes
LinkedIn, so your account stays safe. It just kills the boring copy-typing.

Usage:
    # paste a single post, end with Ctrl-D (mac/linux) or Ctrl-Z then Enter (win)
    python linkedin_post_extractor.py

    # parse one saved post file
    python linkedin_post_extractor.py post.txt

    # batch: parse every .txt post in a folder
    python linkedin_post_extractor.py posts_folder/

Output goes to leads.xlsx in the current folder (change OUTPUT_XLSX below).
"""

from __future__ import annotations
import os, re, sys, glob
from datetime import datetime

# ---------- config ----------
OUTPUT_XLSX = "leads.xlsx"
COLUMNS = [
    "date_added", "contact_person", "title", "company",
    "primary_email", "all_emails", "phone",
    "service_type", "linkedin_url", "company_url", "post_age", "hashtags",
]

# ---------- patterns ----------
EMAIL_RE   = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
PROFILE_RE = re.compile(r"https?://(?:[a-z]{2,3}\.)?linkedin\.com/in/[A-Za-z0-9\-_%]+/?", re.I)
COMPANY_RE = re.compile(r"https?://(?:[a-z]{2,3}\.)?linkedin\.com/company/[A-Za-z0-9\-_%]+/?", re.I)
MD_LINK_RE = re.compile(r"\[([^\]]*)\]\((mailto:[^)]+|https?://[^)]+)\)")
HASHTAG_RE = re.compile(r"#(\w+)")
AGE_RE     = re.compile(r"\b(\d+)\s*(h|hr|hrs|hour|hours|d|day|days|w|wk|week|weeks|mo|month|months|yr|year|years)\b", re.I)
# Indian mobiles + generic international, while avoiding long id numbers in URLs
PHONE_RE   = re.compile(r"(?:\+?91[\-\s]?)?\b[6-9]\d{9}\b|\+\d{1,3}[\-\s]?\d{3}[\-\s]?\d{3,8}")
NOISE_LINK = re.compile(
    r"^[\s•·]*$|•|^(?:1st|2nd|3rd|\d+(?:st|nd|rd|th))\+?$|"
    r"^(?:follow|connect|message|see more|feed post|view post|"
    r"view profile|like|comment|share|repost|reactions?|comments?)$",
    re.I,
)
# A string that looks like a real person's name (letters, spaces, dots, hyphens)
_LOOKS_LIKE_NAME = re.compile(r"^[A-Za-z][A-Za-z\s\.\-']{1,50}$")

SERVICE_MAP = [
    (re.compile(r"contract\s*staffing|contractstaffing", re.I),                       "Contract Staffing"),
    (re.compile(r"\bit\s*staffing|itstaffing", re.I),                                 "IT Staffing"),
    (re.compile(r"bench\s*sales|benchsales|\bbench\b", re.I),                          "Bench Sales"),
    (re.compile(r"vendor|recruitment\s*partner|recruitmentpartner", re.I),            "Vendor Partnership"),
    (re.compile(r"it\s*recruitment|itrecruitment|recruitmentagency|staffing\s*solutions", re.I), "IT Recruitment"),
    (re.compile(r"talent\s*acquisition|talentacquisition", re.I),                     "Talent Acquisition"),
]

GENERIC_DOMAINS = {"gmail.com", "yahoo.com", "outlook.com", "hotmail.com",
                   "rediffmail.com", "icloud.com", "ymail.com", "live.com"}


# ---------- core ----------
def _clean_url(u: str) -> str:
    return u.split("?")[0].rstrip("/")


def parse_linkedin_post(text: str) -> dict:
    lines = [ln.strip() for ln in text.splitlines()]
    md_links = MD_LINK_RE.findall(text)  # [(text, url), ...]

    # --- emails (dedup, keep original casing) ---
    emails, seen = [], set()
    for e in EMAIL_RE.findall(text):
        if e.lower() not in seen:
            seen.add(e.lower())
            emails.append(e)

    # --- author: first /in/ link with a real person-like name ---
    author_name, author_url = "", ""
    for txt, url in md_links:
        if "/in/" in url.lower():
            t = txt.strip()
            if t and not NOISE_LINK.search(t) and _LOOKS_LIKE_NAME.match(t):
                author_name, author_url = t, _clean_url(url)
                break
    if not author_url:
        m = PROFILE_RE.search(text)
        if m:
            author_url = _clean_url(m.group(0))
    # Try to derive name from URL slug when link text wasn't a real name
    # e.g. "sanyam-billa-aaa75b280" → "Sanyam Billa"
    if not author_name and author_url:
        slug = author_url.rstrip("/").split("/in/")[-1]
        parts = [p.capitalize() for p in slug.split("-") if p.isalpha() and len(p) > 1]
        if parts:
            author_name = " ".join(parts)
    if not author_name:  # plain-text paste fallback: first real name-like line
        for ln in lines:
            if ln and not ln.startswith(("[", "#")) and "http" not in ln and "@" not in ln:
                if not AGE_RE.fullmatch(ln.replace("•", "").strip()):
                    if not NOISE_LINK.search(ln.strip()) and _LOOKS_LIKE_NAME.match(ln.strip()):
                        author_name = ln
                        break

    # --- title: the line just after the author, before the timestamp ---
    title, start = "", 0
    for i, ln in enumerate(lines):
        if author_name and author_name in ln:
            start = i
            break
    for s in lines[start + 1:start + 6]:
        if not s or s.startswith(("[", "#")) or "http" in s.lower():
            continue
        if s.lower() in ("follow", "connect", "message") or "@" in s:
            continue
        if AGE_RE.search(s) and "•" in s:
            continue
        if AGE_RE.fullmatch(s.replace("•", "").strip()):
            continue
        if len(s) <= 60:
            title = s
            break

    # --- company: /company/ link → card headline lines → email domain ---
    company, company_url = "", ""
    for txt, url in md_links:
        if "/company/" in url.lower():
            company, company_url = txt.strip(), _clean_url(url)
            break
    if not company_url:
        m = COMPANY_RE.search(text)
        if m:
            company_url = _clean_url(m.group(0))
    if not company and author_name:
        # LinkedIn card layout: Name → [Title at/|/@ Company] → post text
        # Scan the 2-5 lines after the author name for a company hint
        _at_re   = re.compile(r"\bat\b", re.I)
        _pipe_re = re.compile(r"\s[|@]\s")
        try:
            idx = next(i for i, l in enumerate(lines) if author_name in l)
            for ln in lines[idx + 1: idx + 6]:
                if not ln or NOISE_LINK.search(ln) or "http" in ln or "@" in ln:
                    continue
                if AGE_RE.search(ln):
                    continue
                if ln == author_name:  # skip repeated author name (card text duplication)
                    continue
                # "Senior Recruiter at Flair Consulting" → take after "at"
                if _at_re.search(ln):
                    company = _at_re.split(ln, maxsplit=1)[-1].strip()
                    if len(company) < 60:
                        break
                    company = ""
                # "Senior Recruiter | Flair Consulting" → take after "|"
                elif _pipe_re.search(ln):
                    company = _pipe_re.split(ln)[-1].strip()
                    if len(company) < 60:
                        break
                    company = ""
                # Or use the whole line if it looks like a company name
                elif 3 < len(ln) < 60 and not any(c.isdigit() for c in ln[:4]):
                    company = ln
                    break
        except StopIteration:
            pass
    if not company and emails:
        for e in emails:
            dom = e.split("@")[1].lower()
            if dom not in GENERIC_DOMAINS:
                company = dom.split(".")[0].capitalize()
                break

    # --- phones ---
    phones, pseen = [], set()
    for p in PHONE_RE.findall(text):
        digits = re.sub(r"\D", "", p)
        if 10 <= len(digits) <= 13 and digits not in pseen:
            pseen.add(digits)
            phones.append(p.strip())
    phone = "; ".join(phones)

    # --- primary email: prefer one matching the author's first name ---
    primary = ""
    if emails:
        primary = emails[0]
        if author_name:
            fn = re.split(r"\s+", author_name.strip())[0].lower()
            for e in emails:
                if e.split("@")[0].lower().startswith(fn):
                    primary = e
                    break

    # --- hashtags + service type ---
    tags = []
    for h in HASHTAG_RE.findall(text):
        if h.lower() not in {t.lower() for t in tags}:
            tags.append(h)
    services = [label for rx, label in SERVICE_MAP if rx.search(text)]
    service_type = " / ".join(dict.fromkeys(services))  # unique, keep order

    # --- post age ---
    m = AGE_RE.search(text)
    age = m.group(0).strip() if m else ""

    return {
        "date_added": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "contact_person": author_name,
        "title": title,
        "company": company,
        "primary_email": primary,
        "all_emails": "; ".join(emails),
        "phone": phone,
        "service_type": service_type,
        "linkedin_url": author_url,
        "company_url": company_url,
        "post_age": age,
        "hashtags": ", ".join(tags),
    }


# ---------- outputs ----------
def append_to_xlsx(parsed: dict, path: str = OUTPUT_XLSX) -> bool:
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        print("Install openpyxl first:  pip install openpyxl --break-system-packages")
        return False

    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "leads"
        ws.append(COLUMNS)

    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(COLUMNS, row))
        if d.get("primary_email"):
            existing.add(("email", str(d["primary_email"]).lower()))
        if d.get("linkedin_url"):
            existing.add(("url", str(d["linkedin_url"]).lower()))

    ke = ("email", parsed["primary_email"].lower()) if parsed["primary_email"] else None
    ku = ("url", parsed["linkedin_url"].lower()) if parsed["linkedin_url"] else None
    if (ke and ke in existing) or (ku and ku in existing):
        print(f"- skipped duplicate: {parsed['contact_person']} ({parsed['primary_email']})")
        return False

    ws.append([parsed.get(c, "") for c in COLUMNS])
    wb.save(path)
    print(f"+ added: {parsed['contact_person']} | {parsed['company']} | {parsed['primary_email']}")
    return True


def append_to_gsheet(parsed: dict, sheet_name="Leads", worksheet="Sheet1",
                     creds_json="service_account.json") -> bool:
    """Append to a Google Sheet instead of / in addition to Excel.
    Setup: pip install gspread  ->  put a Google service-account JSON next to
    this script as service_account.json  ->  share your sheet with the
    service account's email. Then uncomment the call in main()."""
    import gspread
    gc = gspread.service_account(filename=creds_json)
    ws = gc.open(sheet_name).worksheet(worksheet)
    if not ws.row_values(1):
        ws.append_row(COLUMNS)
    col = COLUMNS.index("primary_email") + 1
    existing = [e.lower() for e in ws.col_values(col)[1:]]
    if parsed["primary_email"] and parsed["primary_email"].lower() in existing:
        print(f"- skipped duplicate in Sheet: {parsed['primary_email']}")
        return False
    ws.append_row([parsed.get(c, "") for c in COLUMNS])
    print(f"+ added to Google Sheet: {parsed['contact_person']}")
    return True


# ---------- entry point ----------
def _iter_inputs(arg):
    if arg is None:
        print("Paste the LinkedIn post, then press Ctrl-D:\n")
        yield sys.stdin.read()
    elif os.path.isdir(arg):
        for f in sorted(glob.glob(os.path.join(arg, "*.txt"))):
            with open(f, encoding="utf-8") as fh:
                yield fh.read()
    else:
        with open(arg, encoding="utf-8") as fh:
            yield fh.read()


def main():
    arg = sys.argv[1] if len(sys.argv) > 1 else None
    added = 0
    for raw in _iter_inputs(arg):
        if not raw.strip():
            continue
        parsed = parse_linkedin_post(raw)
        if not any([parsed["primary_email"], parsed["linkedin_url"], parsed["contact_person"]]):
            print("! couldn't find anything useful in that text")
            continue
        if append_to_xlsx(parsed):
            added += 1
        # append_to_gsheet(parsed)   # <- uncomment to also push to Google Sheets
    print(f"\nDone. {added} new lead(s) saved to {OUTPUT_XLSX}.")


if __name__ == "__main__":
    main()
